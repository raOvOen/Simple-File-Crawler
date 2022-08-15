import os
import codecs
import tkinter.messagebox
import docx
import tkinter as tk
import threading
import tkinter.filedialog as fd
import tkinter.ttk as ttk
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

stopped = False
res_adress = []
res_info = []
info_up = []
info_down = []

def block_widgets(state):
    if state:
        btn_start["state"] = tk.DISABLED
        btn_stop["state"] = tk.NORMAL
        btn_ch_dir["state"] = tk.DISABLED
        btn_add_keyword["state"] = tk.DISABLED
        btn_clear["state"] = tk.DISABLED
        entry_add_keyword["state"] = tk.DISABLED
        entry_dir_path["state"] = tk.DISABLED
        entry_file_type["state"] = tk.DISABLED
        checkbox_reg_check["state"] = tk.DISABLED
        checkbox_check_excel["state"] = tk.DISABLED
        checkbox_check_word["state"] = tk.DISABLED
        checkbox_check_name["state"] = tk.DISABLED
        checkbox_check_like_txt["state"] = tk.DISABLED
    else:
        btn_start["state"] = tk.NORMAL
        btn_stop["state"] = tk.DISABLED
        btn_ch_dir["state"] = tk.NORMAL
        btn_add_keyword["state"] = tk.NORMAL
        btn_clear["state"] = tk.NORMAL
        entry_add_keyword["state"] = tk.NORMAL
        entry_dir_path["state"] = tk.NORMAL
        entry_file_type["state"] = tk.NORMAL
        checkbox_reg_check["state"] = tk.NORMAL
        checkbox_check_excel["state"] = tk.NORMAL
        checkbox_check_word["state"] = tk.NORMAL
        checkbox_check_name["state"] = tk.NORMAL
        checkbox_check_like_txt["state"] = tk.NORMAL

def btn_ch_dir_command():
    directory = fd.askdirectory(title="Choose directory", initialdir="/")
    if directory:
        entry_dir_path.delete(0,tk.END)
        entry_dir_path.insert(0,directory)

def btn_add_keyword_command():
    if entry_add_keyword.get() != "":
        listbox_keywords.insert(tk.END,entry_add_keyword.get())
        entry_add_keyword.delete(0, tk.END)

def btn_start_command():
    global stopped
    stopped = False
    block_widgets(True)
    before_start_clear()
    th = threading.Thread(target=thread_start_command)
    th.daemon = True
    th.start()

def btn_stop_command():
    global stopped
    stopped = True

def btn_cp_results_command():
    if listbox_result.size() > 0:
        label_cp_path["text"] = "All results were successfully copied!"
        label_cp_path["justify"]=tk.LEFT
        label_cp_path["fg"] = 'purple'
        result_of_scan = ""
        for i,text in enumerate(listbox_result.get(0, tk.END)):
            result_of_scan = result_of_scan + str(text) + "\n"
        result_of_scan = result_of_scan.rstrip()
        root.clipboard_clear()
        root.clipboard_append(result_of_scan)
        root.update()
        label_cp_path.place(x=240, y=546, height=30)

def btn_clear_command():
    before_start_clear()
    listbox_keywords.delete(0,tk.END)
    label_mentioned["text"] = "Ready to Start!"
    label_cp_path["text"] = ""

def checkbox_check_like_txt_command():
    if check_like_txt.get() == True:
        entry_file_type["state"] = tk.NORMAL
    else:
        entry_file_type["state"] = tk.DISABLED

def listbox_del(event):
    if btn_add_keyword["state"] == tk.NORMAL:
        selected = list(listbox_keywords.curselection())
        selected.reverse()
        for i in selected:
            listbox_keywords.delete(i)

def listbox_open(event):
    if listbox_result.size() > 0:
        temp_str = listbox_result.selection_get()
        if temp_str[0][0] == '[':
            temp_pos = res_info.index(temp_str)
            os.startfile(res_adress[temp_pos])

def listbox_copy_path(event):
    if listbox_result.size() > 0:
        temp_str = listbox_result.selection_get()
        if temp_str[0][0] == '[':
            label_cp_path["text"] = "The selected path was successfully copied!"
            label_cp_path["justify"] = tk.LEFT
            label_cp_path["fg"] = 'green'
            temp_pos = res_info.index(temp_str)
            root.clipboard_clear()
            root.clipboard_append(res_adress[temp_pos])
            root.update()
            label_cp_path.place(x=240, y=546, height=30)

def before_start_clear():
    listbox_result.delete(0, tk.END)
    global res_adress
    res_adress = []
    global res_info
    res_info = []
    global info_up
    info_up = []
    global info_down
    info_down = []

def entry_search_command(event):
    val = event.widget.get()
    if val == '':
        data = info_up + res_info + info_down
    else:
        data = []
        for item in res_info:
            if val.lower() in item.lower():
                data.append(item)

    search_update(data)

def search_update(data):
    listbox_result.delete(0, 'end')
    for item in data:
        listbox_result.insert('end', item)

def btn_help_command():
    tkinter.messagebox.showinfo(title="About",message="Simple File Crawler\n"+
                                                      "\n"+
                                                      "This program was designed to search for the required information (keywords) in Word, Excel and other documents and files.\n"+
                                                      "\n"+
                                                      "How to use the program:\n"+
                                                      "1. Specify or select, using the corresponding button, the root directory of the scan.\n"+
                                                      "2. Using a special window, specify the necessary keywords which will be further searched in the documents. If necessary, you can set auxiliary flags to add phrases of different case, check files and directories names, check Word or Excel documents and some txt like documents. If any keyword was added unintentionally - you can delete it by double-clicking it.\n"+
                                                      "3. Clicking the \"Start\" button will start the scanning process. After its completion, you can select the desired item to copy it or open it - by double clicking on it.\n"+
                                                      "If necessary, you can copy all the results of the scan.\n"+
                                                      "\n"+
                                                      "Made by raOvOen")

def thread_start_command():
    label_mentioned["text"] = "Scanning..."
    progress_bar_search = ttk.Progressbar(root, orient="horizontal", mode="determinate", value=0)
    path = os.walk(entry_dir_path.get())
    cnt_dirs = count_dirs(path)
    progress_bar_search["maximum"] = cnt_dirs
    progress_bar_search.place(x=80, y=556, height=20)
    ins_text = listbox_keywords.get(0,tk.END)
    global stopped
    text = []
    global res_adress
    res_adress = []
    global info_up
    info_up = []
    for i in ins_text:
        text.append(i)
    path = os.walk(entry_dir_path.get())
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    temp_str = f'[{current_time}] Start scanning'
    info_up.append(temp_str)
    listbox_result.insert(tk.END, temp_str)
    temp_str = 'Keywords: '+str(text)
    info_up.append(temp_str)
    listbox_result.insert(tk.END, temp_str)
    for adress, dirs, files in path:
        if check_name.get():
            lookfor_names(adress, dirs, files, text)
        for file in files:
            if not stopped:
                lookfor_text(adress, file, check_word.get(), check_excel.get(), check_like_txt.get(), text)
                progress_bar_search['value'] += 1
    now = datetime.now()
    current_end_time = now.strftime("%H:%M:%S")
    if not stopped:
        temp_str = f'[{current_end_time}] Scanning Finished'
        listbox_result.insert(tk.END,temp_str)
        info_down.append(temp_str)
    else:
        temp_str = f'[{current_end_time}] Scanning Stopped'
        listbox_result.insert(tk.END, temp_str)
        info_down.append(temp_str)
    progress_bar_search.stop()
    progress_bar_search.destroy()
    label_mentioned["text"] = f"The keywords are mentioned {len(res_adress)} time(s)!"
    block_widgets(False)

def count_dirs(path):
    counter = 0
    for adress, dirs, files in path:
        counter += len(files)
    return counter

def lookfor_names(adr, dirs, files, text):
    for dir in dirs:
        for keyword in text:
            if check_reg.get():
                if keyword.lower() in dir.lower():
                    full_adr = adr + os.sep + dir
                    temp_str = '[Name] Mentioned ' +  keyword + ' in ' + '[' + full_adr + ']'
                    listbox_result.insert(tk.END, temp_str)
                    res_adress.append((full_adr))
                    res_info.append(temp_str)
            else:
                if keyword in dir.lower():
                    full_adr = adr + os.sep + dir
                    temp_str = '[Name] Mentioned ' +  keyword + ' in ' + '[' + full_adr + ']'
                    listbox_result.insert(tk.END, temp_str)
                    res_adress.append((full_adr))
                    res_info.append(temp_str)
    for file in files:
        for keyword in text:
            if check_reg.get():
                if keyword.lower() in file.lower():
                    full_adr = adr + os.sep + file
                    temp_str = '[Name] Mentioned ' + keyword + ' in ' + '[' + full_adr + ']'
                    listbox_result.insert(tk.END, temp_str)
                    res_adress.append((full_adr))
                    res_info.append(temp_str)
            else:
                if keyword in file.lower():
                    full_adr = adr + os.sep + file
                    temp_str = '[Name] Mentioned ' + keyword + ' in ' + '[' + full_adr + ']'
                    listbox_result.insert(tk.END, temp_str)
                    res_adress.append((full_adr))
                    res_info.append(temp_str)

def lookfor_text(adr, filename, word, excel, like_txt, text):
    file_ext = os.path.splitext(filename)
    if word:
        if (file_ext[-1] == ".docx" or file_ext[-1] == ".doc"):
            count = 0
            full_adr = adr + os.sep + filename
            try:
                doc_file = docx.Document(full_adr)
                all_paras = doc_file.paragraphs
                act_keyword = "|"
                for para in all_paras:
                    for keyword in text:
                        if check_reg.get():
                            if keyword.lower() in para.text.lower():
                                count += 1
                                if keyword not in act_keyword:
                                    act_keyword += keyword + "|"
                        else:
                            if keyword in para.text:
                                count += 1
                                if keyword not in act_keyword:
                                    act_keyword += keyword + "|"
            except (Exception):
                pass
            finally:
                if count > 0:
                    temp_str = '[Text] Mentioned: ' + act_keyword + ' for ' + str(count) + ' times! [' + full_adr + ']'
                    if entry_search.get != "" and entry_search.get() in temp_str:
                        listbox_result.insert(tk.END, temp_str)
                    res_adress.append(full_adr)
                    res_info.append(temp_str)
    if excel:
        if (file_ext[-1] == ".xlsx" or file_ext[-1] == ".xls"):
            count = 0
            full_adr = adr + os.sep + filename
            try:
                act_keyword = "|"
                wb = load_workbook(full_adr)
                all_sheet_names = wb.sheetnames
                for sheet_name in all_sheet_names:
                    sheet = wb[sheet_name]
                    for i in range (sheet.max_column):
                        for j in range (sheet.max_row):
                            temp_value= str(sheet[get_column_letter(i+1)+str(j+1)].value)
                            for keyword in text:
                                if check_reg.get():
                                    if keyword.lower() in temp_value.lower():
                                        count += 1
                                        if keyword not in act_keyword:
                                            act_keyword += keyword + "|"
                                else:
                                    if keyword in temp_value.lower():
                                        count += 1
                                        if keyword not in act_keyword:
                                            act_keyword += keyword + "|"
            except Exception as e:
                pass
            finally:
                if count > 0:
                    temp_str = '[Text] Mentioned: ' + act_keyword + ' for ' + str(count) + ' times! [' + full_adr + ']'
                    if entry_search.get != "" and entry_search.get() in temp_str:
                        listbox_result.insert(tk.END, temp_str)
                    res_adress.append(full_adr)
                    res_info.append(temp_str)
    if like_txt:
        full_ext = entry_file_type.get().split(' ')
        for ext in full_ext:
            if (file_ext[-1] == ext):
                count = 0
                full_adr = adr + os.sep + filename
                try:
                    act_keyword = "|"
                    with codecs.open(full_adr, 'r', 'UTF-8') as temp_file:
                        for line in temp_file:
                            for keyword in text:
                                if check_reg.get():
                                    if keyword.lower() in line.rstrip().lower():
                                        count += 1
                                        if keyword not in act_keyword:
                                            act_keyword += keyword + "|"
                                else:
                                    if keyword in line.rstrip():
                                        count += 1
                                        if keyword not in act_keyword:
                                            act_keyword += keyword + "|"
                except Exception as e:
                    pass
                finally:
                    if count > 0:
                        temp_str = '[Text] Mentioned: ' + act_keyword + ' for ' + str(count) + ' times! [' + full_adr + ']'
                        if entry_search.get != "" and entry_search.get() in temp_str:
                            listbox_result.insert(tk.END, temp_str)
                        res_adress.append(full_adr)
                        res_info.append(temp_str)

if __name__ == '__main__':
    root = tk.Tk()
    root.title("Simple File Crawler by raOvOen")
    root.resizable(width=False, height=False)
    window_width = 800
    window_height = 600
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_cordinate = int((screen_width / 2) - (window_width / 2))
    y_cordinate = int((screen_height / 2) - (window_height / 2))
    root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))

    btn_ch_dir = tk.Button(root)
    btn_ch_dir["text"] = "Choose directory"
    btn_ch_dir.place(x=10, y=20, width=117, height=30)
    btn_ch_dir["command"] = btn_ch_dir_command

    entry_dir_path = tk.Entry(root)
    entry_dir_path["text"] = "Dir"
    entry_dir_path.place(x=140, y=20, width=600, height=30)

    listbox_keywords = tk.Listbox(root)
    listbox_keywords.place(x=320, y=60, width=470, height=120)
    listbox_keywords.bind('<Double-Button-1>', listbox_del)

    btn_add_keyword = tk.Button(root)
    btn_add_keyword["text"] = "Add keyword"
    btn_add_keyword.place(x=140, y=105, width=170, height=35)
    btn_add_keyword["command"] = btn_add_keyword_command

    btn_clear = tk.Button(root)
    btn_clear["text"] = "Clear all"
    btn_clear.place(x=230, y=148, width=80, height=32)
    btn_clear["command"] = btn_clear_command

    entry_add_keyword = tk.Entry(root)
    entry_add_keyword["text"] = "Entry"
    entry_add_keyword.place(x=140, y=63, width=170, height=32)

    btn_start = tk.Button(root)
    btn_start["text"] = "Start"
    btn_start.place(x=140, y=182, width=170, height=40)
    btn_start["command"] = btn_start_command

    btn_stop = tk.Button(root)
    btn_stop["text"] = "Stop"
    btn_stop.place(x=140, y=148, width=80, height=32)
    btn_stop["state"] = tk.DISABLED
    btn_stop["command"] = btn_stop_command

    check_reg = tk.BooleanVar()
    check_reg.set(True)
    checkbox_reg_check = tk.Checkbutton(root)
    checkbox_reg_check["text"] = "Simple Reg check"
    checkbox_reg_check.place(x=0, y=55, width=140, height=30)
    checkbox_reg_check["variable"] = check_reg
    checkbox_reg_check["offvalue"] = False
    checkbox_reg_check["onvalue"] = True

    check_name = tk.BooleanVar()
    check_name.set(True)
    checkbox_check_name = tk.Checkbutton(root)
    checkbox_check_name["text"] = "Check names"
    checkbox_check_name.place(x=0, y=80, width=118, height=30)
    checkbox_check_name["variable"] = check_name
    checkbox_check_name["offvalue"] = False
    checkbox_check_name["onvalue"] = True

    check_word = tk.BooleanVar()
    check_word.set(True)
    checkbox_check_word = tk.Checkbutton(root)
    checkbox_check_word["text"] = "Check word"
    checkbox_check_word.place(x=0, y=105, width=110, height=30)
    checkbox_check_word["variable"] = check_word
    checkbox_check_word["offvalue"] = False
    checkbox_check_word["onvalue"] = True

    check_excel = tk.BooleanVar()
    check_excel.set(False)
    checkbox_check_excel = tk.Checkbutton(root)
    checkbox_check_excel["text"] = "Check excel"
    checkbox_check_excel.place(x=0, y=130, width=110, height=30)
    checkbox_check_excel["variable"] = check_excel
    checkbox_check_excel["offvalue"] = False
    checkbox_check_excel["onvalue"] = True

    check_like_txt = tk.BooleanVar()
    check_like_txt.set(False)
    checkbox_check_like_txt = tk.Checkbutton(root)
    checkbox_check_like_txt["text"] = "Check as TXT:"
    checkbox_check_like_txt.place(x=0, y=155, width=118, height=30)
    checkbox_check_like_txt["variable"] = check_like_txt
    checkbox_check_like_txt["command"] = checkbox_check_like_txt_command
    checkbox_check_like_txt["offvalue"] = False
    checkbox_check_like_txt["onvalue"] = True

    entry_file_type = tk.Entry(root)
    entry_file_type["text"] = "EntryType"
    entry_file_type.insert(tk.END, ".txt .json .xml .py")
    entry_file_type["state"] = tk.DISABLED
    entry_file_type.place(x=12, y=190, width=118, height=30)

    scrollbar = tk.Scrollbar(root, orient="horizontal")
    scrollbar.pack(side="bottom",fill="x")
    listbox_result = tk.Listbox(root,xscrollcommand=scrollbar.set,height=20)
    listbox_result.place(x=10, y=230, width=780, height=310)
    listbox_result.bind('<Double-Button-1>', listbox_open)
    listbox_result.bind('<<ListboxSelect>>',listbox_copy_path)
    scrollbar.config(command=listbox_result.xview)

    btn_help = tk.Button(root)
    btn_help["text"] = "?"
    btn_help.place(x=750, y=10, width=40, height=40)
    btn_help["command"] = btn_help_command

    btn_cp_results = tk.Button(root)
    btn_cp_results["text"] = "Copy all results"
    btn_cp_results.place(x=670, y=546, width=120, height=30)
    btn_cp_results["command"] = btn_cp_results_command

    label_mentioned = tk.Label(root)
    label_mentioned["text"] = "Ready to Start!"
    label_mentioned["justify"]=tk.LEFT
    label_mentioned.place(x=10, y=546, height=30)

    label_cp_path = tk.Label(root)

    entry_search = tk.Entry(root)
    entry_search.place(x=375, y=190, width=415, height=30)
    entry_search.bind('<KeyRelease>', entry_search_command)

    label_search = tk.Label(root)
    label_search["text"] = "Search:"
    label_search["justify"]=tk.LEFT
    label_search.place(x=325, y=190, height=30)

    stopped = False
    root.mainloop()
